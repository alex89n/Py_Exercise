from openpyxl import *
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Fill
import sys
import os

RFQ_template = load_workbook("RFQ_Template.xlsx")
RFQ_sheet = RFQ_template.active
RFQ_save = "RFQ.xlsx"

def pretabaj(r, razmak, document, ime_ponudjaca = ""):
    red = []
    for row in document['A' + str(r):'N' + str(r)]:
        for cell in row:
            # print cell.coordinate, cell.value   
            red.append(cell.value)
    if ime_ponudjaca != "":
        red[0] = ime_ponudjaca
    print red
    print "-----------------------------------------------"

    for i in range(1,15):
        if ime_ponudjaca == "":        
            RFQ_sheet.cell(row=razmak,column=i).font = Font(bold=True)
        else:
            RFQ_sheet.cell(row=razmak,column=i).font = Font(bold=False)
            
        RFQ_sheet.cell(row=razmak,column=i).value = red[i-1]

        
def ponuda(br_komponenti, br_ponudjaca, redni_broj_ponudjaca, ponuda, ime_ponudjaca = ""):
    for i in range(1, br_komponenti):
        pretabaj(i+1,(redni_broj_ponudjaca+1)+(i-1)*br_ponudjaca, ponuda, ime_ponudjaca)

        
def main():
    try:    
        file = sys.argv[1]
        with open(file, 'r') as file:
            line = file.readlines()
        br_komponenti = int(sys.argv[2])
    except:
        print "-----------------------------------------------"
        print "\nERROR - Morate uneti ispravne parametre"
        print "<ime_skripte> <spisak_ponudjaca> <br_komponenti>\n"
        return
    else:
        # with open(file, 'r') as file:
            # line = file.readlines()
        
        # provera ispravnosti fajla sa imenima ponudjaca
        for i in range(0,len(line)):
            if line[i].find("_") < 0:
                print "-----------------------------------------------"
                print "\nERROR - Neispravan format fajla sa imenima ponudjaca"
                return
    
    for i in range(0,len(line)):        
        if str(line[i][len(line[i])-1]) == "\n":
            wb = load_workbook(str(line[i][0:-1]))
        else:
            wb = load_workbook(str(line[i]))
    
        sheet = wb.active
        ime =  line[i].find("_")
        ponuda(br_komponenti+1, len(line)+1+2, i+1, sheet, line[i][0:ime])
        # len(line)+1 - ovo +1 je razmak (broj redova) izmedju pojedninih komponenti.
    
    try:
        RFQ_template.save(RFQ_save) # sacuvaj dokument
    except:
        print "\nERROR - Zatvorite " + RFQ_save + " dokument\n"
        return
    else:
        print "\nSuccess - Dokument " + RFQ_save + " kreiran\n"
        os.startfile("RFQ.xlsx")
   
   
if __name__ == "__main__":
    main()
    
    