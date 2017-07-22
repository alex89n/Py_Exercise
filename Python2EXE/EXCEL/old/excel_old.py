from openpyxl import *
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Fill
import os
    
wb1 = load_workbook("VITA_RK2017_01_15_OBLO_MPQ1_RFQ.xlsx")
wb2 = load_workbook("RICH_Quotations from RICHTEK COM SERENA CHAN.xlsx")
wb3 = load_workbook("ALANT_Offers for Milos (2)-dopuna.xlsx")
wb4 = load_workbook("TED_RK2017_01_15_OBLO_MPQ1_RFQ - ponuda.xlsx")
wb5 = load_workbook("IC_KopijaRK2017_01_17_OBLO_MPQ1_RFQ 20 01 17.xlsx")
wb6 = load_workbook("VITA2_Ponuda- RK2017_01_15_OBLO_MPQ1_RFQ.xlsx")

sheet1 = wb1.active
sheet2 = wb2.active
sheet3 = wb3.active
sheet4 = wb4.active
sheet5 = wb5.active
sheet6 = wb6.active

Original =load_workbook("_RK2017_01_17_OBLO_MPQ1_RFQ.xlsx") 
RFQ = load_workbook("RFQ_Template.xlsx")

Original_sheet = Original.active
RFQ_sheet = RFQ.active


def pretabaj(r, razmak, document, ime_ponudjaca = ""):
    red = []
    for row in document['A' + str(r):'N' + str(r)]:
        for cell in row:    
            print cell.coordinate, cell.value   
            red.append(cell.value)
    if ime_ponudjaca != "":
        red[0] = ime_ponudjaca
    print red
    print "----------------------------"

    for i in range(1,15):
        if ime_ponudjaca == "":        
            RFQ_sheet.cell(row=razmak,column=i).font = Font(bold=True)
        else:
            RFQ_sheet.cell(row=razmak,column=i).font = Font(bold=False)
        RFQ_sheet.cell(row=razmak,column=i).value = red[i-1]

       

def ponuda(br_komponenti, br_ponudjaca, redni_broj_ponudjaca, ponuda, ime_ponudjaca = ""):
    for i in range(1, br_komponenti):
        pretabaj(i+1,(redni_broj_ponudjaca+1)+(i-1)*br_ponudjaca, ponuda, ime_ponudjaca)


    
ponuda(15, 11, 1, Original_sheet)
ponuda(15, 11, 2, sheet1, "VITA")
# ponuda(15, 11, 3, sheet2, "RICH")
# ponuda(15, 11, 4, sheet3, "ALANT")
# ponuda(15, 11, 5, sheet4, "TED")
# ponuda(15, 11, 6, sheet5, "IC")
# ponuda(15, 11, 7, sheet6, "VITA2")
            
# for i in range(1, 68):
    # if i == 1:
        # pretabaj_2(i+1, i+1, Original_sheet)
    # else:
        # pretabaj_2(i+1,2+(i-1)*8, Original_sheet)
        
# for i in range(1, 9):
    # if i == 1:
        # pretabaj(i+2, i+2, sheet1, "VITA")
    # else:
        # pretabaj(i+1,3+(i-1)*5, sheet1, "VITA")
        
# for i in range(1, 9):
    # if i == 1:
        # pretabaj(i+3, i+3, sheet2, "ATL")
    # else:
        # pretabaj(i+1,4+(i-1)*5, sheet2, "ATL")
        
RFQ.save("RFQ_old.xlsx")
os.startfile("RFQ_old.xlsx")
    
    
    
    
    # red = []
# ws2 = wb.create_sheet("Drugi")
# n = 5
# for i in xrange(2,5):
    # for row in ws.iter_rows(min_row=i, max_col=3, max_row=i):
        # for cell in row:
            # print(cell)
            # red.append(cell.value)
    # print red


    # for row in ws2.iter_rows(min_row=n, max_col=3, max_row=n):
        # ws2.append(red)
  
        # # for cell in row:
            # # cell.value.append(red)
    # red = []


# from openpyxl import Workbook
# from openpyxl.compat import range
# from openpyxl.utils import get_column_letter

# wb = Workbook()

# dest_filename = 'empty_book.xlsx'

# ws1 = wb.active
# ws1.title = "range names"

# for row in range(1, 40):
    # ws1.append(range(600))

# ws2 = wb.create_sheet(title="Pi")

# ws2['F5'] = 3.14

# ws3 = wb.create_sheet(title="Data")


# wb.save(filename = dest_filename)

# wb.save("S2.xlsx")


# from openpyxl import load_workbook
# wb = load_workbook(filename = 'empty_book.xlsx')
# sheet_ranges = wb['range names']
# print sheet_ranges
# print(sheet_ranges['D18'].value)




